"""
Agentic Markets — Generate XLSX Database
=========================================
Scarica 5 anni di dati storici da football-data.org e genera un XLSX
professionale con tutte le tab strutturate.

Il file viene salvato in: data/agentic_markets_db.xlsx

Poi viene uploadato su Google Drive via MCP.
"""
import os, sys, time, json, requests
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict

# ── openpyxl imports ─────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Config ────────────────────────────────────────────────────────────────────
FDORG_API_KEY = os.getenv("FOOTBALL_DATA_ORG_API_KEY", "")
FDORG_BASE    = "https://api.football-data.org/v4"
SEASONS       = [2020, 2021, 2022, 2023, 2024]
LEAGUES = {
    "PL":  "Premier League",
    "SA":  "Serie A",
    "PD":  "La Liga",
    "BL1": "Bundesliga",
    "FL1": "Ligue 1",
    "CL":  "Champions League",
    "EL":  "Europa League",
}

OUTPUT_DIR  = Path(__file__).parent.parent / "data"
OUTPUT_FILE = OUTPUT_DIR / "agentic_markets_db.xlsx"

# ── Palette colori ────────────────────────────────────────────────────────────
DARK_BG    = "0B0E14"
CYAN       = "00E5FF"
GREEN      = "32E0A0"
AMBER      = "FFB547"
VIOLET     = "9B8CFF"
RED        = "FF5D6C"
WHITE      = "E6EDF7"
MUTED      = "7D8AA3"
PANEL      = "0E131C"
PANEL2     = "121826"
LINE       = "1A2233"
LINE2      = "243049"

TAB_COLORS = {
    "README":            "0E1520",
    "Match_Results":     "0A1C15",
    "Team_Stats":        "0A1420",
    "League_Stats":      "12152A",
    "Odds_Calibration":  "201710",
    "Predictions_Log":   "0D1E1E",
    "Bets_Log":          "1A0E0E",
    "PnL_Monthly":       "0D1F12",
    "Model_Performance": "1A1A0C",
    "Context_Module":    "161020",
}

HDR_FILL  = PatternFill("solid", fgColor=PANEL2)
HDR_FONT  = Font(bold=True, color=CYAN, size=10, name="JetBrains Mono" if False else "Consolas")
CELL_FONT = Font(color=WHITE, size=9)
MUTED_FONT= Font(color=MUTED, size=9)

def h_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def h_font(hex_color: str, bold=False, sz=9) -> Font:
    return Font(color=hex_color, bold=bold, size=sz)

def thin_border() -> Border:
    s = Side(style="thin", color=LINE2)
    return Border(bottom=s)

def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header_row(ws, headers: list[str], accent: str = CYAN):
    ws.append(headers)
    row = ws.max_row
    for col in range(1, len(headers)+1):
        c = ws.cell(row=row, column=col)
        c.fill  = h_fill(PANEL2)
        c.font  = Font(bold=True, color=accent, size=9, name="Consolas")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

def write_data_row(ws, values: list, accent: str = WHITE, alt: bool = False):
    ws.append(values)
    row = ws.max_row
    bg = "101820" if alt else PANEL
    for col in range(1, len(values)+1):
        c = ws.cell(row=row, column=col)
        c.fill  = h_fill(bg)
        c.font  = Font(color=accent, size=9)
        c.alignment = Alignment(vertical="center")
        c.border = thin_border()

def style_sheet(ws, tab_color: str):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False

def freeze_and_autofilter(ws, last_col: str):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}1"

# ── football-data.org fetch ───────────────────────────────────────────────────

def fdorg_get(endpoint: str, params=None, retries=3):
    url = f"{FDORG_BASE}{endpoint}"
    headers = {"X-Auth-Token": FDORG_API_KEY}
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=15)
            if r.status_code == 429:
                wait = int(r.headers.get("X-RequestCounter-Reset", 65))
                print(f"    rate limit — attendo {wait}s …")
                time.sleep(wait)
                continue
            if r.status_code == 200:
                return r.json()
            print(f"    HTTP {r.status_code} per {endpoint}")
            return None
        except Exception as e:
            print(f"    Errore: {e}")
            if attempt < retries-1: time.sleep(5)
    return None

def fetch_all_matches() -> list[dict]:
    all_matches = []
    total = len(LEAGUES) * len(SEASONS)
    n = 0
    t0 = time.time()

    for league in LEAGUES:
        for season in SEASONS:
            n += 1
            print(f"  [{n}/{total}] {league} {season} …", end=" ", flush=True)
            data = fdorg_get(f"/competitions/{league}/matches",
                             {"season": season, "status": "FINISHED"})
            if data:
                ms = []
                for m in data.get("matches", []):
                    sc = m.get("score",{}).get("fullTime",{})
                    hg = sc.get("home"); ag = sc.get("away")
                    if hg is None or ag is None: continue
                    res = "H" if hg > ag else "A" if ag > hg else "D"
                    ms.append({
                        "match_id":    str(m.get("id","")),
                        "date":        (m.get("utcDate",""))[:10],
                        "season":      season,
                        "league":      league,
                        "league_name": LEAGUES[league],
                        "home_team":   (m.get("homeTeam",{}).get("shortName")
                                        or m.get("homeTeam",{}).get("name","")),
                        "away_team":   (m.get("awayTeam",{}).get("shortName")
                                        or m.get("awayTeam",{}).get("name","")),
                        "home_goals":  int(hg),
                        "away_goals":  int(ag),
                        "result":      res,
                        "matchday":    m.get("matchday",""),
                        "stage":       m.get("stage","REGULAR_SEASON"),
                    })
                print(f"{len(ms)} match")
                all_matches.extend(ms)
            else:
                print("0 match (no data)")

            # rate limit: 10 req/min free tier → sleep 6.5s
            elapsed = time.time() - t0
            if n % 10 == 0 and elapsed < 62:
                wait = 65 - elapsed
                print(f"    ⏳ finestra rate limit — attendo {wait:.0f}s …")
                time.sleep(wait)
                t0 = time.time()
            else:
                time.sleep(6.5)

    return all_matches

def compute_team_stats(matches):
    stats = defaultdict(lambda: {"gp":0,"w":0,"d":0,"l":0,"gf":0,"ga":0,"pts":0})
    for m in matches:
        hk = (m["season"], m["league"], m["home_team"])
        ak = (m["season"], m["league"], m["away_team"])
        hg, ag = m["home_goals"], m["away_goals"]
        for k, gf, ga in [(hk, hg, ag), (ak, ag, hg)]:
            stats[k]["gp"] += 1; stats[k]["gf"] += gf; stats[k]["ga"] += ga
        if hg > ag:
            stats[hk]["w"]+=1; stats[hk]["pts"]+=3; stats[ak]["l"]+=1
        elif ag > hg:
            stats[ak]["w"]+=1; stats[ak]["pts"]+=3; stats[hk]["l"]+=1
        else:
            stats[hk]["d"]+=1; stats[hk]["pts"]+=1; stats[ak]["d"]+=1; stats[ak]["pts"]+=1
    return [
        (s, l, LEAGUES.get(l,l), t,
         v["gp"], v["w"], v["d"], v["l"],
         v["gf"], v["ga"], v["gf"]-v["ga"], v["pts"])
        for (s,l,t), v in sorted(stats.items())
    ]

def compute_league_stats(matches):
    buckets = defaultdict(list)
    for m in matches:
        buckets[(m["season"], m["league"])].append(m)
    rows = []
    for (s, l), ms in sorted(buckets.items()):
        n = len(ms)
        if n == 0: continue
        tg = sum(m["home_goals"]+m["away_goals"] for m in ms)
        hg = sum(m["home_goals"] for m in ms)
        ag = sum(m["away_goals"] for m in ms)
        hw = sum(1 for m in ms if m["result"]=="H")
        dw = sum(1 for m in ms if m["result"]=="D")
        aw = sum(1 for m in ms if m["result"]=="A")
        rows.append((
            s, l, LEAGUES.get(l,l), n,
            round(tg/n,3), round(hg/n,3), round(ag/n,3),
            round(hw/n*100,1), round(dw/n*100,1), round(aw/n*100,1),
            "1" if l in ("PL","SA","PD","BL1","FL1") else "2",
            datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        ))
    return rows

# ── Sheet builders ────────────────────────────────────────────────────────────

def build_readme(wb: Workbook):
    ws = wb.create_sheet("README")
    style_sheet(ws, "0A0F1A")

    data = [
        ["AGENTIC MARKETS — PREDICTION DATABASE v5.0"],
        [""],
        ["Ultimo aggiornamento", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["Spreadsheet ID", "1aFbrx-w2uI4tRvHjEfWskNcfMuWY5GgwCYNMYE3Q8Mo"],
        ["Dashboard live", "https://agentic-markets-roan.vercel.app"],
        [""],
        ["TAB", "DESCRIZIONE", "FONTE", "AGGIORNAMENTO"],
        ["Match_Results",    "Risultati storici 2020-2024",                "football-data.org",    "settimanale"],
        ["Team_Stats",       "Statistiche per squadra×stagione",           "football-data.org",    "settimanale"],
        ["League_Stats",     "Metriche campionato×stagione",               "football-data.org",    "settimanale"],
        ["Odds_Calibration", "Quote storiche per CLV",                     "The Odds API / Betfair","live"],
        ["Predictions_Log",  "Previsioni Dixon-Coles",                     "DB Neon",              "ogni ora"],
        ["Bets_Log",         "Storico scommesse paper + live",             "DB Neon",              "live"],
        ["PnL_Monthly",      "P&L per mese e campionato",                  "calcolato da Bets_Log","live"],
        ["Model_Performance","Brier, hit rate, CLV per campionato",        "calcolato",            "settimanale"],
        ["Context_Module",   "Match type factors e league tier",           "Context Module v5.0",  "statico"],
        [""],
        ["CAMPIONATI COPERTI"],
        ["PL",  "Premier League (Inghilterra)"],
        ["SA",  "Serie A (Italia)"],
        ["PD",  "La Liga (Spagna)"],
        ["BL1", "Bundesliga (Germania)"],
        ["FL1", "Ligue 1 (Francia)"],
        ["CL",  "Champions League"],
        ["EL",  "Europa League"],
        [""],
        ["STAGIONI", "2020, 2021, 2022, 2023, 2024"],
        [""],
        ["PIPELINE v5.0"],
        ["DataCollector → ModelAgent → ContextService v5.0 → AnalystAgent → StrategistAgent → RiskManagerAgent → TraderAgent"],
        [""],
        ["CONTEXT MODULE v5.0"],
        ["LeagueStrengthAnalyzer",      "Tier 1-5 basato su efficienza mercato + liquidità"],
        ["LeagueOddsProfiler",          "Anomaly detection quote (z-score σ=1.5)"],
        ["LeaguePredictabilityTracker", "Hit rate + CLV; auto-suspend se CLV < 0"],
        ["MatchTypeClassifier",         "10 tipi: DERBY, TITLE_DECIDER, RELEGATION, SHORT_REST…"],
        ["CompetitionTypeFactors",      "Penalità stake: DERBY -30%, DEAD_RUBBER -50%, STANDARD ±0%"],
    ]

    ws.append([""])  # row 1 spacer
    for row in data:
        ws.append(row)

    # Title style
    ws["A2"].font  = Font(bold=True, color=CYAN, size=16)
    ws["A2"].fill  = h_fill("060A10")
    ws.row_dimensions[2].height = 32

    # Section headers
    for cell_ref, color in [("A7", CYAN), ("A17", GREEN), ("A25", AMBER), ("A27", VIOLET), ("A30", GREEN)]:
        try:
            ws[cell_ref].font = Font(bold=True, color=color, size=10)
        except Exception:
            pass

    # Tab list rows
    for row_idx in range(8, 17):
        for col in [1,2,3,4]:
            c = ws.cell(row=row_idx, column=col)
            c.fill = h_fill(PANEL)
            c.font = Font(color=WHITE if col==1 else MUTED, size=9)
            c.border = thin_border()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18

def build_match_results(wb: Workbook, matches: list[dict]):
    ws = wb.create_sheet("Match_Results")
    style_sheet(ws, "0A1C15")
    ws.sheet_view.showGridLines = False

    headers = [
        "match_id","date","season","league","league_name",
        "home_team","away_team","home_goals","away_goals","result",
        "home_xg","away_xg","home_shots","away_shots",
        "matchday","stage",
    ]
    write_header_row(ws, headers, CYAN)

    accent_map = {"H": GREEN, "D": AMBER, "A": RED}
    for i, m in enumerate(sorted(matches, key=lambda x:(x["season"],x["league"],x["date"]))):
        res = m.get("result","")
        row = [m.get(h,"") for h in headers]
        write_data_row(ws, row, alt=i%2==0)
        # Color the result cell
        result_col = headers.index("result") + 1
        ws.cell(ws.max_row, result_col).font = Font(
            color=accent_map.get(res, WHITE), bold=True, size=9
        )

    freeze_and_autofilter(ws, "P")
    set_col_widths(ws, [10,11,7,5,18,22,22,6,6,7,7,7,7,7,8,18])
    ws.row_dimensions[1].height = 20

def build_team_stats(wb: Workbook, matches: list[dict]):
    ws = wb.create_sheet("Team_Stats")
    style_sheet(ws, "0A1420")

    headers = [
        "season","league","league_name","team",
        "gp","wins","draws","losses",
        "goals_for","goals_against","goal_diff","points",
        "xg","xga","xgd","ppda_att","ppda_def","form_last5",
    ]
    write_header_row(ws, headers, GREEN)

    for i, row in enumerate(compute_team_stats(matches)):
        write_data_row(ws, list(row) + ["","","","",""], alt=i%2==0)
        # Colorize goal_diff
        gd_col = 11
        gd = row[10]
        ws.cell(ws.max_row, gd_col).font = Font(
            color=GREEN if gd > 0 else (RED if gd < 0 else WHITE), size=9
        )

    freeze_and_autofilter(ws, "R")
    set_col_widths(ws, [7,5,18,22,5,5,5,5,6,6,6,6,7,7,7,7,7,12])

def build_league_stats(wb: Workbook, matches: list[dict]):
    ws = wb.create_sheet("League_Stats")
    style_sheet(ws, "12152A")

    headers = [
        "season","league","league_name","matches_played",
        "avg_goals","avg_home_goals","avg_away_goals",
        "home_win_%","draw_%","away_win_%",
        "avg_xg_home","avg_xg_away","predictability",
        "league_tier","last_updated",
    ]
    write_header_row(ws, headers, VIOLET)

    for i, row in enumerate(compute_league_stats(matches)):
        write_data_row(ws, list(row), alt=i%2==0)

    freeze_and_autofilter(ws, "O")
    set_col_widths(ws, [7,5,18,8,8,9,9,9,7,9,10,10,12,8,12])

def build_odds_calibration(wb: Workbook):
    ws = wb.create_sheet("Odds_Calibration")
    style_sheet(ws, "201710")

    headers = [
        "match_id","date","league","league_name",
        "home_team","away_team","bookmaker",
        "opening_home","opening_draw","opening_away",
        "closing_home","closing_draw","closing_away",
        "market_margin_%","clv_home","clv_draw","clv_away","source",
    ]
    write_header_row(ws, headers, AMBER)

    # Note rows
    notes = [
        ["# Popolato live dall'Agent System quando vengono registrate scommesse."],
        ["# CLV = Closing Line Value. CLV > 0 = bet piazzata a valore migliore del mercato di chiusura."],
        ["# Fonte principale: The Odds API (Betfair Exchange closing odds)."],
    ]
    for note in notes:
        ws.append(note)
        ws.cell(ws.max_row, 1).font = Font(color=MUTED, size=8, italic=True)

    set_col_widths(ws, [10,11,5,18,22,22,15,9,9,9,9,9,9,10,8,8,8,15])

def build_predictions_log(wb: Workbook):
    ws = wb.create_sheet("Predictions_Log")
    style_sheet(ws, "0D1E1E")

    headers = [
        "match_id","league","league_name","home_team","away_team","kickoff",
        "p_home","p_draw","p_away","lambda_home","lambda_away",
        "odds_home","odds_draw","odds_away",
        "edge","best_selection","model_matches","computed_at",
    ]
    write_header_row(ws, headers, CYAN)

    notes = [
        ["# Popolato ogni ora dal sistema. Edge > 3% = value bet candidata."],
        ["# p_home + p_draw + p_away = 1.0 (modello Dixon-Coles calibrato)."],
    ]
    for note in notes:
        ws.append(note)
        ws.cell(ws.max_row, 1).font = Font(color=MUTED, size=8, italic=True)

    freeze_and_autofilter(ws, "R")
    set_col_widths(ws, [10,5,18,22,22,16,7,7,7,8,8,8,8,8,8,12,8,16])

def build_bets_log(wb: Workbook):
    ws = wb.create_sheet("Bets_Log")
    style_sheet(ws, "1A0E0E")

    headers = [
        "bet_id","placed_at","match_id","league","league_name",
        "home_team","away_team","kickoff",
        "selection","odds","stake","status","mode","profit_loss",
    ]
    write_header_row(ws, headers, RED)

    notes = [
        ["# Popolato live. mode = PAPER | LIVE. status = won | lost | pending | voided."],
        ["# profit_loss: positivo = profitto, negativo = perdita."],
    ]
    for note in notes:
        ws.append(note)
        ws.cell(ws.max_row, 1).font = Font(color=MUTED, size=8, italic=True)

    freeze_and_autofilter(ws, "N")
    set_col_widths(ws, [7,16,10,5,18,22,22,16,10,7,7,9,7,10])

def build_pnl_monthly(wb: Workbook):
    ws = wb.create_sheet("PnL_Monthly")
    style_sheet(ws, "0D1F12")

    headers = [
        "month","bets","won","lost","pending",
        "win_rate_%","gross_pnl_eur","roi_%",
        "avg_odds","total_staked","leagues",
    ]
    write_header_row(ws, headers, GREEN)

    notes = [
        ["# Aggregato mensile calcolato da Bets_Log. ROI = gross_pnl / total_staked."],
        ["# Aggiornato automaticamente dall'agent system."],
    ]
    for note in notes:
        ws.append(note)
        ws.cell(ws.max_row, 1).font = Font(color=MUTED, size=8, italic=True)

    set_col_widths(ws, [9,6,5,5,7,9,11,7,8,11,20])

def build_model_performance(wb: Workbook):
    ws = wb.create_sheet("Model_Performance")
    style_sheet(ws, "1A1A0C")

    headers = [
        "league","league_name","bets_evaluated","won","hit_rate_%",
        "gross_pnl_eur","avg_edge_%","brier_score","clv_avg",
        "value_bets_pct","model_version","notes",
    ]
    write_header_row(ws, headers, AMBER)

    placeholders = [
        ["PL",  "Premier League",  0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
        ["SA",  "Serie A",         0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
        ["PD",  "La Liga",         0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
        ["BL1", "Bundesliga",      0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
        ["FL1", "Ligue 1",         0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
        ["CL",  "Champions League",0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
        ["EL",  "Europa League",   0, 0, 0.0, 0.0, "", "", "", "", "Dixon-Coles v5.0", ""],
    ]
    for i, row in enumerate(placeholders):
        write_data_row(ws, row, alt=i%2==0)

    freeze_and_autofilter(ws, "L")
    set_col_widths(ws, [5,18,10,6,9,11,9,10,8,11,16,20])

def build_context_module(wb: Workbook):
    ws = wb.create_sheet("Context_Module")
    style_sheet(ws, "161020")

    # SECTION 1: Match Types
    ws.append(["MATCH TYPE FACTORS — Context Module v5.0"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color=VIOLET, size=12)
    ws.row_dimensions[ws.max_row].height = 24
    ws.append([""])

    mt_headers = [
        "match_type","description",
        "stake_multiplier","confidence_penalty",
        "typical_leagues","auto_skip","notes",
    ]
    write_header_row(ws, mt_headers, VIOLET)

    match_types = [
        ["DERBY_NATIONAL",   "Derby nazionale/cittadino",                   0.70, -0.10, "Tutti",    "No", "Alta variabilità emotiva"],
        ["DERBY_REGIONAL",   "Derby regionale",                             0.80, -0.08, "Tutti",    "No", ""],
        ["TITLE_DECIDER",    "Match decisivo per il titolo",                0.85, -0.05, "Top 5",    "No", "Alta attenzione mediatica"],
        ["RELEGATION",       "Scontro diretto salvezza",                    0.80, -0.08, "Tutti",    "No", "Alta motivazione"],
        ["DEAD_RUBBER",      "Partita ininfluente fine stagione",           0.50, -0.15, "Tutti",    "No", "Rotazioni probabili"],
        ["CUP_SPILLOVER",    "Effetto coppa / campo neutro",                0.75, -0.10, "CL/EL",   "No", ""],
        ["NEUTRAL_VENUE",    "Campo neutro (finale gara unica)",            0.75, -0.08, "CL/EL",   "No", ""],
        ["SHORT_REST",       "Squadra con <72h dall'ultimo match",          0.75, -0.08, "Tutti",    "No", ""],
        ["EUROPEAN_HANGOVER","Dopo gara europea infrasettimanale",          0.80, -0.06, "Tutti",    "No", ""],
        ["ROTATION",         "Rotazioni attese (coppa/stanchezza)",         0.60, -0.12, "Tutti",    "No", ""],
        ["STANDARD",         "Match standard senza fattori speciali",       1.00,  0.00, "Tutti",    "No", ""],
    ]
    for i, row in enumerate(match_types):
        write_data_row(ws, row, alt=i%2==0)
        mult_col = 3
        mult = row[2]
        ws.cell(ws.max_row, mult_col).font = Font(
            color=GREEN if mult >= 1.0 else (AMBER if mult >= 0.75 else RED),
            bold=True, size=9
        )

    ws.append([""])

    # SECTION 2: League Tiers
    ws.append(["LEAGUE TIER CONFIGURATION"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color=CYAN, size=12)
    ws.row_dimensions[ws.max_row].height = 24
    ws.append([""])

    lt_headers = [
        "league","league_name","tier",
        "avg_market_efficiency","liquidity_score",
        "recommended_edge_min_%","notes",
    ]
    write_header_row(ws, lt_headers, CYAN)

    league_tiers = [
        ["PL",  "Premier League",  1, 0.97, 0.95, 2.0, "Top 1 per liquidità Betfair"],
        ["SA",  "Serie A",         1, 0.95, 0.88, 2.0, ""],
        ["PD",  "La Liga",         1, 0.96, 0.90, 2.0, ""],
        ["BL1", "Bundesliga",      1, 0.95, 0.87, 2.0, ""],
        ["FL1", "Ligue 1",         1, 0.93, 0.82, 2.5, "Meno liquido dei top 4"],
        ["CL",  "Champions League",2, 0.96, 0.85, 2.5, "Alta efficienza pre-match"],
        ["EL",  "Europa League",   2, 0.90, 0.70, 3.0, "Meno coperto dai bookmakers"],
    ]
    for i, row in enumerate(league_tiers):
        write_data_row(ws, row, alt=i%2==0)
        tier_col = 3
        tier = row[2]
        ws.cell(ws.max_row, tier_col).font = Font(
            color=GREEN if tier == 1 else AMBER, bold=True, size=9
        )

    set_col_widths(ws, [22, 25, 8, 18, 14, 20, 35])

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("📡 Download dati da football-data.org …")
    print(f"   API key: {FDORG_API_KEY[:8]}…")
    print(f"   Leghe: {', '.join(LEAGUES.keys())}")
    print(f"   Stagioni: {SEASONS}")
    print()

    all_matches = fetch_all_matches()
    print(f"\n✅ {len(all_matches)} match totali scaricati")

    # Salva cache JSON
    cache_path = OUTPUT_DIR / "matches_cache.json"
    cache_path.write_text(json.dumps(all_matches, ensure_ascii=False, indent=2))
    print(f"💾 Cache salvata: {cache_path}")

    print("\n📊 Costruzione XLSX …")
    wb = Workbook()

    # Rimuovi sheet default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_readme(wb)
    build_match_results(wb, all_matches)
    build_team_stats(wb, all_matches)
    build_league_stats(wb, all_matches)
    build_odds_calibration(wb)
    build_predictions_log(wb)
    build_bets_log(wb)
    build_pnl_monthly(wb)
    build_model_performance(wb)
    build_context_module(wb)

    wb.save(OUTPUT_FILE)
    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"\n✅ XLSX salvato: {OUTPUT_FILE} ({size_kb} KB)")
    print(f"   Righe Match_Results: {len(all_matches)}")
    print()
    print("📌 Prossimo step: uploadare su Google Drive via MCP")

if __name__ == "__main__":
    main()
