"""Tennis edge/CLV backtest — does our surface-Elo beat the tennis market?

Self-contained: tennis-data.co.uk gives results + Pinnacle closing odds (PSW/PSL)
+ ranks in one file per year (ATP and WTA), the tennis analog of football-data.co.uk.
No name-join to Sackmann needed. Files cached under data/tennis_data_uk/.

Walk-forward, no leakage:
  - players are assigned to slot A / slot B by alphabetical name order — a rule
    INDEPENDENT of who won — so the Elo prediction never sees the outcome;
  - the pre-match Pinnacle prices are simply re-attached to their slot;
  - Elo is updated only AFTER the match is scored.

Measures, on the eval slice:
  - accuracy: model vs market-favourite (lower odds)
  - Brier: model vs de-vigged market (calibration)
  - ROI/CLV: flat-stake model value picks priced at Pinnacle closing, bootstrap CI.

Run:  .venv/bin/python -m scripts.backtest_tennis_edge
"""
from __future__ import annotations

import sys
import urllib.request
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from models.elo_surface import EloSurfaceModel  # noqa: E402

CACHE = ROOT / "data" / "tennis_data_uk"
YEARS = [2021, 2022, 2023, 2024]
WARMUP_YEARS = [2021]           # Elo burn-in only
CALIB_YEARS = [2022]            # fit isotonic calibration (Elo keeps updating)
EVAL_YEARS = [2023, 2024]       # honest out-of-sample eval
EDGE_BUCKETS = [(0.0, 0.05), (0.05, 0.10), (0.10, 0.20), (0.20, 1.0)]
N_BOOT = 2000
RNG_SEED = 20260701
UA = {"User-Agent": "Mozilla/5.0 (agentic-markets tennis data loader)"}


def _url(tour: str, year: int) -> str:
    # ATP: /2024/2024.xlsx ; WTA: /2024w/2024.xlsx
    suffix = "" if tour == "ATP" else "w"
    return f"http://www.tennis-data.co.uk/{year}{suffix}/{year}.xlsx"


def load_year(tour: str, year: int) -> list[dict]:
    CACHE.mkdir(parents=True, exist_ok=True)
    fp = CACHE / f"{tour.lower()}_{year}.xlsx"
    if not fp.exists():
        req = urllib.request.Request(_url(tour, year), headers=UA)
        last = None
        for attempt in range(4):
            try:
                with urllib.request.urlopen(req, timeout=60) as r:  # noqa: S310 (trusted host)
                    fp.write_bytes(r.read())
                last = None
                break
            except Exception as e:  # noqa: BLE001 — transient network, retry
                last = e
        if last is not None:
            print(f"  ! skip {tour} {year}: {last!r}")
            return []
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    col = {name: i for i, name in enumerate(header)}
    need = ("Date", "Surface", "Winner", "Loser")
    if not all(n in col for n in need):
        wb.close()
        return []
    rows = []
    for r in it:
        try:
            date = r[col["Date"]]
            surface = (r[col["Surface"]] or "").strip().lower() if r[col["Surface"]] else "hard"
            winner = (r[col["Winner"]] or "").strip()
            loser = (r[col["Loser"]] or "").strip()
            if not winner or not loser or date is None:
                continue

            def num(key):
                if key in col and r[col[key]] not in (None, ""):
                    try:
                        return float(r[col[key]])
                    except (TypeError, ValueError):
                        return None
                return None

            # Pinnacle closing; fall back to Bet365 then market average
            psw = num("PSW") or num("B365W") or num("AvgW")
            psl = num("PSL") or num("B365L") or num("AvgL")
            # best price available across books (line-shopping) — where real edge lives
            maxw = num("MaxW") or psw
            maxl = num("MaxL") or psl
            rows.append({
                "date": date, "tour": tour, "surface": surface,
                "winner": winner, "loser": loser,
                "wrank": num("WRank"), "lrank": num("LRank"),
                "odds_w": psw, "odds_l": psl,
                "max_w": maxw, "max_l": maxl,
            })
        except Exception:  # noqa: BLE001 — skip malformed row
            continue
    wb.close()
    return rows


def brier_2way(p_a: float, a_won: int) -> float:
    return (p_a - a_won) ** 2 + ((1 - p_a) - (1 - a_won)) ** 2


def roi_ci(stakes: np.ndarray, profits: np.ndarray, rng: np.random.Generator):
    total = stakes.sum()
    if total <= 0:
        return float("nan"), (float("nan"), float("nan")), 0
    roi = profits.sum() / total * 100
    n = len(stakes)
    boot = []
    for _ in range(N_BOOT):
        idx = rng.integers(0, n, n)
        s = stakes[idx].sum()
        if s > 0:
            boot.append(profits[idx].sum() / s * 100)
    lo, hi = np.percentile(boot, [2.5, 97.5]) if boot else (float("nan"), float("nan"))
    return roi, (float(lo), float(hi)), int((stakes > 0).sum())


def _year(d) -> int:
    return d.year if hasattr(d, "year") else 0


def run() -> None:
    from sklearn.isotonic import IsotonicRegression

    rng = np.random.default_rng(RNG_SEED)
    all_rows: list[dict] = []
    for tour in ("ATP", "WTA"):
        for yr in YEARS:
            rows = load_year(tour, yr)
            all_rows.extend(rows)
            print(f"  {tour} {yr}: {len(rows)} matches")
    all_rows.sort(key=lambda r: (r["date"], r["tour"]))
    print(f"Total: {len(all_rows)} matches")

    elo = {"ATP": EloSurfaceModel(), "WTA": EloSurfaceModel()}
    calib_x: dict[str, list] = {"ATP": [], "WTA": []}
    calib_y: dict[str, list] = {"ATP": [], "WTA": []}
    eval_rows: list[dict] = []

    # single walk-forward pass: predict raw p, stash for calibration or eval, then update Elo
    for r in all_rows:
        tour = r["tour"]
        m = elo[tour]
        yr = _year(r["date"])
        if yr not in WARMUP_YEARS:
            a, b = sorted([r["winner"], r["loser"]])          # slot rule independent of outcome
            a_won = 1 if a == r["winner"] else 0
            p_raw = m.predict(a, b, r["surface"])["p1"]
            if yr in CALIB_YEARS:
                calib_x[tour].append(p_raw); calib_y[tour].append(a_won)
            elif yr in EVAL_YEARS:
                eval_rows.append({**r, "p_raw": p_raw, "a_won": a_won,
                                  "odds_a": r["odds_w"] if a_won else r["odds_l"],
                                  "odds_b": r["odds_l"] if a_won else r["odds_w"],
                                  "max_a": r["max_w"] if a_won else r["max_l"],
                                  "max_b": r["max_l"] if a_won else r["max_w"]})
        m.update(r["winner"], r["loser"], r["surface"])

    # fit isotonic calibration per tour on the CALIB slice
    iso = {}
    for tour in ("ATP", "WTA"):
        if len(calib_x[tour]) > 200:
            ir = IsotonicRegression(out_of_bounds="clip", y_min=0.0, y_max=1.0)
            ir.fit(calib_x[tour], calib_y[tour])
            iso[tour] = ir
    print(f"\nCalibration fit on {CALIB_YEARS} | eval on {EVAL_YEARS} "
          f"({len(eval_rows)} matches)")

    # ── calibration quality (raw vs calibrated vs market) ──
    def brier_list(rows, key):
        return float(np.mean([brier_2way(r[key], r["a_won"]) for r in rows])) if rows else float("nan")

    for r in eval_rows:
        ir = iso.get(r["tour"])
        r["p_cal"] = float(ir.predict([r["p_raw"]])[0]) if ir else r["p_raw"]
        if r["odds_a"] and r["odds_b"] and r["odds_a"] > 1 and r["odds_b"] > 1:
            ia, ib = 1 / r["odds_a"], 1 / r["odds_b"]
            r["p_mkt"] = ia / (ia + ib)
        else:
            r["p_mkt"] = None

    priced = [r for r in eval_rows if r["p_mkt"] is not None]
    print("\n" + "=" * 92)
    print("CALIBRATION (2-way Brier, lower=better) — eval slice")
    print("=" * 92)
    print(f"  Elo RAW        : {brier_list(priced, 'p_raw'):.5f}")
    print(f"  Elo CALIBRATED : {brier_list(priced, 'p_cal'):.5f}")
    print(f"  Market         : {brier_list(priced, 'p_mkt'):.5f}")
    macc = np.mean([(r['p_cal'] >= 0.5) == (r['a_won'] == 1) for r in priced])
    kacc = np.mean([(r['p_mkt'] >= 0.5) == (r['a_won'] == 1) for r in priced])
    print(f"  Accuracy  model {macc:.4f}  vs  market {kacc:.4f}  (n={len(priced)})")

    # market-anchored blend = the shippable Track-A product (serve better than raw Elo)
    print("  -- market-anchored blend  p = w*mkt + (1-w)*elo_cal --")
    best = None
    for w in (0.3, 0.5, 0.7, 0.85):
        b = float(np.mean([brier_2way(w * r["p_mkt"] + (1 - w) * r["p_cal"], r["a_won"]) for r in priced]))
        print(f"     w={w:.2f}: Brier {b:.5f}")
        if best is None or b < best[1]:
            best = (w, b)
    print(f"  best blend w={best[0]:.2f} -> {best[1]:.5f}  (raw Elo {brier_list(priced,'p_raw'):.5f}, market {brier_list(priced,'p_mkt'):.5f})")

    # ── selective edge by bucket, priced at Pinnacle-close AND best-available (Max) ──
    def bucket_roi(rows, price_key_a, price_key_b, lo, hi):
        stakes, profits = [], []
        for r in rows:
            for p, odds, won in ((r["p_cal"], r[price_key_a], r["a_won"]),
                                 (1 - r["p_cal"], r[price_key_b], 1 - r["a_won"])):
                if not odds or odds <= 1:
                    continue
                edge = p * odds - 1.0
                if lo <= edge < hi:
                    stakes.append(1.0)
                    profits.append((odds - 1.0) if won else -1.0)
        return np.array(stakes), np.array(profits)

    for price_label, ka, kb in (("Pinnacle CLOSE", "odds_a", "odds_b"),
                                 ("BEST-ODDS (Max)", "max_a", "max_b")):
        print("\n" + "=" * 92)
        print(f"SELECTIVE EDGE @ {price_label} — calibrated picks by edge bucket")
        print("=" * 92)
        print(f"{'edge bucket':<16}{'bets':>7}{'ROI%':>10}{'CI95':>22}{'gate':>7}")
        print("-" * 92)
        for lo, hi in EDGE_BUCKETS:
            st, pr = bucket_roi(priced, ka, kb, lo, hi)
            if len(st) == 0:
                continue
            roi, ci, nb = roi_ci(st, pr, rng)
            gate = "PASS" if ci[0] > 0 else "off"
            label = f"{lo*100:.0f}-{hi*100:.0f}%" if hi < 1 else f">{lo*100:.0f}%"
            print(f"{label:<16}{nb:>7}{roi:>+10.2f}{('[%+.2f,%+.2f]' % ci):>22}{gate:>7}")
    print("=" * 92)
    print("ROI>0 with CI95 excluding 0 = genuine edge. Best-odds = realistic sharp price.")


if __name__ == "__main__":
    run()
